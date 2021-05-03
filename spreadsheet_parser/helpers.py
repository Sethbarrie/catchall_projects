import sys
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename
ALPHA = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"



""" 
Creates dictionary from specified column, 
requires: 
    spreadsheet, 
    column name not letter,
    *optional* starting row, defaults to 2, since openpyxl is 1 indexed.
Iterates row by row until it doesn't find a value. Creates dictionary
with the cell value as the key and the row as the value
"""
def create_column_dictionary(spreadsheet, callback, starting_column = 1):
    return_dictionary = {}
    current_column = starting_column

    end_of_columns = end_of_spreadsheet(spreadsheet[ALPHA[current_column]])
    while not end_of_columns:
        kv_tuple = callback(create_value_tuple(spreadsheet[ALPHA[current_column]]), return_dictionary, current_column)
        return_dictionary[kv_tuple[0]] = kv_tuple[1]
        current_column += 1
        end_of_columns = current_column > 25 or end_of_spreadsheet(spreadsheet[ALPHA[current_column]])
    return return_dictionary

def create_rows_dictionary(spreadsheet, callback, starting_row = 2):
    return_dictionary = {}

    current_row = starting_row
    end_of_rows = end_of_spreadsheet(spreadsheet[current_row])
    while not end_of_rows:
        kv_tuple = callback(create_value_tuple(spreadsheet[current_row]), return_dictionary, current_row)
        return_dictionary[kv_tuple[0]] = kv_tuple[1]
        current_row += 1
        end_of_rows = end_of_spreadsheet(spreadsheet[current_row])
    return return_dictionary

"""
Returns the index from the tuple of the specific value, 
used in conjunction with letter_column_from_tuple
"""
def term_index_from_tuple(header, term):
    for i, cell in enumerate(header):
        if cell.value == term:
            return i
    return -1

"""
Returns the column letter of the specific value
"""
def letter_column_from_tuple(header, term):
    counter = term_index_from_tuple(header, term)
    return ALPHA[counter]


def error_check(array):
    """
    array = [
        (<Value>,<Error message if failed>, <Callback to test value>),
        (1, "Output spreadsheet needs a column named 'Example'", lambda x: x == -1),
        ...
    ]
    """
    for error_container in array:
        if error_container[2](error_container[0]):
            print(error_container[1])
            sys.exit(1)


def manipulate_rows_by_column(spreadsheet, callback, column, starting_row = 1):
    current_row_number = starting_row
    cell_string = f"{column}{current_row_number}"
    current_cell = spreadsheet[cell_string]
    while not current_cell:
        spreadsheet[cell_string] = callback(current_cell)
        current_row_number += 1
        cell_string = f"{column}{current_row_number}"
        current_row_tuple = spreadsheet[cell_string]

def manipulate_columns_by_row(spreadsheet, callback, row, starting_column = "A"):
    current_column_letter = starting_column
    current_cell = spreadsheet[current_column_letter + str(row)]
    cell_string = f"{current_column_letter}{row}"
    while not current_cell:
        spreadsheet[cell_string] = callback(current_cell)
        if current_column_letter == "Z":
            current_cell = false
        else:
            current_column_letter = ALPHA[(ALPHA.index(current_column_letter) + 1)]
            cell_string = f"{current_column_letter}{row}"
            current_row_tuple = spreadsheet[cell_string]

def enumerate_rows(spreadsheet, callback, starting_row = 1):
    current_row_number = starting_row
    end_of_rows = end_of_spreadsheet(spreadsheet[current_row_number])
    while not end_of_rows:
        #This is a function that will enumerate one row and manipulate values that way
        callback(spreadsheet, current_row_number)
        current_row_number += 1
        end_of_rows = end_of_spreadsheet(spreadsheet[current_row_number])

#Since all the tuples return cells, I created this so I can just have values from the cells
def create_value_tuple(cell_tuple):
    mid_array = []
    for cell in cell_tuple:
        mid_array.append(cell.value)
    return tuple(mid_array)

#Considered having this as any, but if there were missing values it wouldn't go to the end of the spreadsheet
def end_of_spreadsheet(row_tuple):
    return all(map(lambda x: x.value is None, row_tuple))

#Runs a callback on the current row, then inserts values from returned tuple
def manipulate_row(spreadsheet, callback, row):
    """
    insert_tuple = (
            (<Column letter to insert at>, <Value to insert in that cell>),
            ("A", "Store name 50"),
            ("B", 25),
            ("C", 125),
            ("D", 55.3),
            ("E", 5),
            ("F", 0.05),
        )
    """
    insert_tuple = callback(create_value_tuple(spreadsheet[row]))
    for val in insert_tuple:
        spreadsheet[f"{val[0]}{str(row)}"] = val[1]

#If there is an ordered column, find and insert an empty row where it would be sorted
def insert_row(spreadsheet, insert_column, insert_value, callback, starting_row = 2):
    current_row = starting_row
    lower_insert_value = 0
    cell_string = f"{insert_column}{current_row}"
    found_insert_spot = spreadsheet[cell_string].value != None
    while found_insert_spot:
        current_value = callback(spreadsheet[cell_string].value)
        if current_value > insert_value and insert_value > lower_insert_value:
            found_insert_spot = false
        else:
            lower_insert_value = current_value
            current_row += 1
            cell_string = f"{insert_column}{current_row}"
            found_insert_spot = spreadsheet[cell_string].value != None
    spreadsheet.insert_rows(current_row)
    return current_row

#Returns either cells value or default value if cell was empty
def set_new_value(check_value, new_value):
    if check_value == None:
        return new_value
    else:
        return check_value



class Worksheets():
    Tk().withdraw()
    input_spreadsheet_path = askopenfilename()
    output_spreadsheet_path = askopenfilename()
    
    error_check([
        (input_spreadsheet_path,"Input spreadsheet wasn't selected!", lambda x: x == ""),
        (output_spreadsheet_path,"Output spreadsheet wasn't selected!", lambda x: x == "")
    ]) 
    
    wb_input = load_workbook(input_spreadsheet_path)
    input_spreadsheet = wb_input.active
    wb_output = load_workbook(output_spreadsheet_path)
    output_spreadsheet = wb_output.active

    def save_work(self):
        #Save your progress!
        self.wb_input.save(self.input_spreadsheet_path)
        self.wb_output.save(self.output_spreadsheet_path)