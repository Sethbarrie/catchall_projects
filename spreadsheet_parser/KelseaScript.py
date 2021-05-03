from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import helpers

def main():

    # Tk().withdraw()
    # input_spreadsheet_path = askopenfilename()
    # output_spreadsheet_path = askopenfilename()

    # helpers.error_check([
    #     (input_spreadsheet_path,"Input spreadsheet wasn't selected!"),
    #     (output_spreadsheet_path,"Output spreadsheet wasn't selected!")
    # ])

    # wb_input = load_workbook(input_spreadsheet_path)
    # input_spreadsheet = wb_input.active

    # wb_output = load_workbook(output_spreadsheet_path)
    # output_spreadsheet = wb_output.active

    input_spreadsheet_path = "./xlsx_files/copy/input.xlsx"
    output_spreadsheet_path = "./xlsx_files/copy/output_empty.xlsx"

    wb_input = load_workbook(input_spreadsheet_path)
    input_spreadsheet = wb_input.active

    wb_output = load_workbook(output_spreadsheet_path)
    output_spreadsheet = wb_output.active

    helpers.error_check([
        (input_spreadsheet, "Input spreadsheet couldn't be read!"),
        (output_spreadsheet, "Output spreadsheet couldn't be read!")
    ])


    store_number_rows = helpers.create_column_dictionary(output_spreadsheet, "Store Number")
    store_payment_outputs = helpers.populate_input_types_per_store(input_spreadsheet)
    new_store_payment_outputs = {}

    store_number_column = helpers.letter_column_from_tuple(output_spreadsheet['1'], "Store Number")
    cash_column = helpers.letter_column_from_tuple(output_spreadsheet['1'], "Cash")
    credit_column = helpers.letter_column_from_tuple(output_spreadsheet['1'], "Credit")
    cashless_column = helpers.letter_column_from_tuple(output_spreadsheet['1'], "Cashless (mobile)")


    helpers.error_check([
        (store_number_column, "Output spreadsheet needs a column named 'Store Number'"), 
        (cash_column, "Output spreadsheet needs a column named 'Cash'"), 
        (credit_column, "Output spreadsheet needs a column named 'Credit'"), 
        (cashless_column, "Output spreadsheet needs a column named 'Cashless (mobile)'")
    ])

    for store_num in store_payment_outputs:
        payment_info = store_payment_outputs[store_num]
        store_row = 0
        if store_num in store_number_rows:
            store_row = str(store_number_rows[store_num])
            
            helpers.update_number_value_in_spreadsheet(
                f"{cash_column}{store_row}", 
                output_spreadsheet, 
                payment_info["Cash"]
            )
            helpers.update_number_value_in_spreadsheet(
                f"{credit_column}{store_row}", 
                output_spreadsheet, 
                payment_info["Credit"]
            )
            helpers.update_number_value_in_spreadsheet(
                f"{cashless_column}{store_row}",
                output_spreadsheet, 
                payment_info["Cashless (mobile)"]
            )
        else:
            new_store_payment_outputs[store_num] = store_payment_outputs[store_num]

    for store_num in new_store_payment_outputs:
        payment_info = store_payment_outputs[store_num]

        helpers.insert_values_into_spreadsheet(
            [store_number_column, cash_column, credit_column, cashless_column],
            output_spreadsheet, 
            payment_info,
            int(store_num)
        )
    def total_function(spreadsheet, row):
        print("In total function")
        for cell in spreadsheet[row]:
            print(cell.value)
        val1 = int(spreadsheet["B" + str(row)].value)
        val2 = int(spreadsheet["C" + str(row)].value)
        val3 = int(spreadsheet["D" + str(row)].value)
        new_total = 0 if spreadsheet["E" + str(row)].value is None else int(spreadsheet["E" + str(row)].value)
        spreadsheet["E" + str(row)] = new_total + (val1 + val2 + val3)
        spreadsheet["F" + str(row)] = (int(spreadsheet["E" + str(row)].value)/ 2)

    helpers.enumerate_rows(output_spreadsheet, total_function, 2)
    wb_input.save(input_spreadsheet_path)
    wb_output.save(output_spreadsheet_path)
    
if __name__ == "__main__":
    main()